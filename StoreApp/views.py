from .models import Customer_info,CustomUser, Draw_date
from django.http import JsonResponse, HttpRequest
from django.shortcuts import render, HttpResponse
from .serializers import Customer_serializer, UserSerializer, Draw_date_serializer
from rest_framework import viewsets
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate
from django.core.exceptions import ObjectDoesNotExist
from openpyxl import Workbook
from rest_framework.views import APIView
from datetime import datetime , timezone
# Create your views here.



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def register_user(request):
    if request.method == 'POST':
        serializer = UserSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



@api_view(['POST'])
def user_login(request):
    if request.method == 'POST':
        username = request.data.get('username')
        password = request.data.get('password')

        user = None
        if '@' in username:
            try:
                user = CustomUser.objects.get(email=username)
            except ObjectDoesNotExist:
                pass

        if user:
            user = authenticate(username=user.username, password=password)
        else:
            user = authenticate(username=username, password=password)

        if user:
            token, _ = Token.objects.get_or_create(user=user)
            return Response({'token': token.key}, status=status.HTTP_200_OK)

        return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def user_logout(request):
    if request.method == 'POST':
        try:
            # Delete the user's token to logout
            request.user.auth_token.delete()
            return Response({'message': 'Successfully logged out.'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['get'])
@permission_classes([IsAuthenticated])
def get_user_info(request):
    user = request.user
    user_info = {
        'username': user.username,
        'email' : user.email,
    }
    return Response(user_info)

# return record which our not soft delete
class CustomerInfo(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = Customer_info.objects.filter(is_deleted=False) 
    serializer_class = Customer_serializer


#####################################################                          ###############################################
##################################################### NON DELETED RECORD EXCEL ################################################
@permission_classes([IsAuthenticated])
def export_to_excel(request):
    # Retrieve date range from frontend
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    if from_date_str and to_date_str:
    # Convert date strings to datetime objects
        from_date = datetime.strptime(from_date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)


        from_date_formatted = from_date.strftime("%d-%m-%Y")
        to_date_formatted = to_date.strftime("%d-%m-%Y")

        customer_data = Customer_info.objects.filter(is_deleted=False, dr_date__range=[from_date, to_date])


        wb = Workbook()
        ws = wb.active

        # Write headers
        headers = ['Name', 'Mobile_No', 'Invoice_No', 'Invoice_Date','Amount', 'Coupon_id', 'draw_date']
        ws.append(headers)

        # Write data to Excel
        for customer_row in customer_data:
            formated_invoice_date = customer_row.Invoice_date.strftime("%d-%m-%Y")
            formated_draw_date = customer_row.dr_date.strftime("%d-%m-%Y")
            ws.append([customer_row.Name, customer_row.Mobile_No, customer_row.Invoice_no,formated_invoice_date, customer_row.Amount, customer_row.Token_id,formated_draw_date])

        # Create a response containing the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Draw_date {from_date_formatted} - {to_date_formatted}.xlsx"'
        wb.save(response)

        return response
    


    # Retrieve all data from the database via DRF view
    customer_data = Customer_info.objects.filter(is_deleted=False) 

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write headers
    headers = ['Name', 'Mobile_No', 'Invoice_No', 'Invoice_Date','Amount', 'Coupon_id', 'draw_date']
    ws.append(headers)

    # Write data to Excel
    for customer_row in customer_data:
        formated_invoice_date = customer_row.Invoice_date.strftime("%d-%m-%Y")
        formated_draw_date = customer_row.dr_date.strftime("%d-%m-%Y")
        ws.append([customer_row.Name, customer_row.Mobile_No, customer_row.Invoice_no,formated_invoice_date, customer_row.Amount, customer_row.Token_id,formated_draw_date])

    # Create a response containing the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="cutomers_data.xlsx"'
    wb.save(response)

    return response



#################################################                      ###################################################
################################################   DELTED RECORD EXCEL ###################################################
@permission_classes([IsAuthenticated])
def export_deleted_to_excel(request):

    # Retrieve date range from frontend
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    

    if from_date_str and to_date_str:
    # Convert date strings to datetime objects
        from_date = datetime.strptime(from_date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)


        from_date_formatted = from_date.strftime("%d-%m-%Y")
        to_date_formatted = to_date.strftime("%d-%m-%Y")

        customer_data = Customer_info.objects.filter(is_deleted=True, dr_date__range=[from_date, to_date])

        # Create an Excel workbook
        wb = Workbook()
        ws = wb.active

        # Write headers
        headers = ['Name', 'Mobile_No', 'Invoice_No', 'Invoice_Date','Amount', 'Coupon_id', 'draw_date', 'IP_Address']
        ws.append(headers)

        # Write data to Excel
        for customer_row in customer_data:
            formated_invoice_date = customer_row.Invoice_date.strftime("%d-%m-%Y")
            formated_draw_date = customer_row.dr_date.strftime("%d-%m-%Y")
            ws.append([customer_row.Name, customer_row.Mobile_No, customer_row.Invoice_no,formated_invoice_date, customer_row.Amount, customer_row.Token_id,formated_draw_date,customer_row.deleted_by_ip])

        # Create a response containing the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Deleted_data Draw_date {from_date_formatted} - {to_date_formatted}.xlsx"'
        wb.save(response)

        return response
    



    # Retrieve all data from the database via DRF view
    customer_data = Customer_info.objects.filter(is_deleted=True) 

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write headers
    headers = ['Name', 'Mobile_No', 'Invoice_No', 'Invoice_Date','Amount', 'Coupon_id', 'draw_date','IP_Address']
    ws.append(headers)

    # Write data to Excel
    for customer_row in customer_data:
        formated_invoice_date = customer_row.Invoice_date.strftime("%d-%m-%Y")
        formated_draw_date = customer_row.dr_date.strftime("%d-%m-%Y")
        ws.append([customer_row.Name, customer_row.Mobile_No, customer_row.Invoice_no,formated_invoice_date, customer_row.Amount, customer_row.Token_id,formated_draw_date,customer_row.deleted_by_ip])

    # Create a response containing the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="deleted_data.xlsx"'
    wb.save(response)

    return response


##################################################################################################################


class DrawDate(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = Draw_date.objects.all()
    serializer_class = Draw_date_serializer



@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def soft_delete_record(request, record_id):

    if request.method == 'PUT':
        record = Customer_info.objects.filter(pk=record_id).first()
        if record:
             # Include both is_deleted and deleted_by_ip from the request data
            data = {'is_deleted': True, 'deleted_by_ip': request.data.get('deleted_by_ip')}
            serializer = Customer_serializer(instance=record, data=data, partial=True, context={'request': request})
            if serializer.is_valid():
                serializer.save()
                return Response({'message': 'Record soft deleted successfully'}, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Record not found'}, status=status.HTTP_404_NOT_FOUND)
        
        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_deleted_records(request):
    deleted_data = Customer_info.objects.filter(is_deleted=True)
    serializer = Customer_serializer(deleted_data, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)
